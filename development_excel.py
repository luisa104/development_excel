# -*- coding: utf-8 -*-
"""
Created on Wed Jul  7 14:06:19 2021

@author: Asus-PC
"""
import pandas as pd
from pandas.core.indexes.api import all_indexes_same
import random
from openpyxl import Workbook
from openpyxl import load_workbook
import os 





def viewColumns():
    choose = int(input("how many columns do you want to see 1-28: "))
    size = len(desired_columns)
    if 1 <= choose <= size:
        size = (size - 1)
        print(date.loc[:,desired_columns[0:choose]])
        
def viewColumns_voice():
    choose = int(input("which column do you want to see: "))
    size = len(desired_columns)
    if 1 <= choose <= size:
        size = (size - 1)
        print(date.loc[:,desired_columns[choose] ] )
        
def columns_purchase():
    choose = int(input("which column do you want to see: "))
    size = len(col)
    if 1 <= choose <= size:
        size = (size - 1)
        print(df.loc[:,col[choose] ] )

def viewColumns_purchase():
    choose = int(input("how many columns do you want to see 1-21: "))
    size = len(col)
    if 1 <= choose <= size:
        size = (size - 1)
        print(df.loc[:,col[0:choose]])


def row_invoice():
    choose = int(input("how many rows do you want to see: "))
    row = date.head(choose)
    print(row)
    
    
def row_single():
    choose = int(input("which row do you want to see: "))
    save = date.iloc[choose] 
    print(save)
    
def row_purchase():
    choose = int(input("how many rows do you want to see: "))
    row = df.head(choose)
    print( row )    
    
def row_single_purchase():
    choose = int(input("which row do you want to see: "))
    save = df.iloc[choose] 
    print(save)
    

def empty():
    op=int(input("which column you want: "))
    count = date.loc[:,col[op]].isna().sum()
    print(f"the empty elements in this column are: {count}")    
    
    
def empty_purchase():
    op=int(input("which column you want: "))
    count = df.loc[:,col[op]].isna().sum()
    print(f"the empty elements in this column are: {count}")   

def attributes():
    op=int(input("which column you want: "))
    value = (date.dtypes.iloc[op])
    print(value)
    
def attributes_purchase():
    op=int(input("which column you want: "))
    value = (df.dtypes.iloc[op])
    print(value)
    
def show_all():
    print(date.dtypes)
    
def show_all_purchase():
    print(df.dtypes)
 
def show_rows():
    number = len(date.index)
    print(number)
    
    
def show_rows_purchase():
    number = len(df.index)
    print(number)
    
    
def show_column():
    number = len(date.columns)
    print(number)
    
    
def show_column_purchase():
    number = len(df.columns)
    print(number)
 
 
def not_null():
    opcion=int(input("which column you want: "))
    count = date.loc[:,col[opcion]].isna().sum()
    B = len(date.loc[:,col[opcion]])
    C = B - count
    print(f"the elemets in this column are:{C} ")


def not_null_purchase():
    opcion=int(input("which column you want: "))
    count = date.loc[:,col[opcion]].isna().sum()
    B = len(date.loc[:,col[opcion]])
    C = B - count
    print(f"the elemets in this column are:{C} ")
    
    
def zise():
    print(date.shape)
    
    
def size_purchase():
    print(df.shape)

def InvoiceDate():
    day = random.randint(10,31)
    month = random.randint(1,9)
    year = random.randint(2020,2020)
    b  = str(day)+"/"+"0"+str(month)+"/"+str(year)
    return  b
        

def POBasedInvoices():
    a =("Invoice Only","Compliant PO","After the Fact PO","Same Day PO")
    var =  random.choice(a)
    return var

def Item():
    ca = "IT-279836","IT-63597","IT-432432"
    n = random.choice(ca)
    return n 




def ItemDescription():
    c = ("6-32 Screw","1/2-13 x 6" ,"Hex Bolt, 18-8 Stainless Steel","1/2-13 Hex Nut")
    b = random.choice(c)
    return b

def DiscountRate():
    val = random.uniform(100,900)
    cor =str(round(val / 100,1)) +str( "%")
    c = cor
    a = c.replace(".",",")
    return a

def SalesTaxRate():
    nu = random.uniform(500,1000)
    nor =str(round(nu / 100,1)) +str( "%")
    c = nor
    a = c.replace(".",",")
    return a

def ShippingFee():
    lu = "50","500","250","75"
    n = random.choice(lu)
    return n 
    
    
def InvoiceStatus():
    lu = "Paid","Open","Overdue"
    n = random.choice(lu)
    return n 
    
def InvoiceType():
    lu = "Ariba Network","Paper"
    n = random.choice(lu)
    return n 
   
def SourceType():

    lu = "Addressable","Unaddressable"
    n = random.choice(lu)
    return n

def SourceSystem():
    lu = "Ariba","Legacy 1","Legacy 2","Legacy 3"
    n = random.choice(lu)
    return n
    
   

def Catalog():
    lu = "Y","N"
    n = random.choice(lu)
    return n
    
def Comments():
    lu ="Revised Contract","No Contract"
    n = random.choice(lu)
    return n
    
"""
count = date.loc[:,col[0]].isna().sum()
B = len(date.loc[:,col[0]])
C = B - count
"""
def null():
     for i in range(1,  (13)):
         if  ws['O' + str(i)].value == None:
             ws['O' + str(i)]  = DiscountRate()
            
     wb2.save(invoice) 
    



 
def null_DiscountAmount():
     for i in range(1,  (13)):
         if  ws['P' + str(i)].value == None:
             ws['P' + str(i)]  =  '=N' + str(i) +'*O' + str(i)
     wb2.save(invoice) 
    



def null_DiscountedPrice():
    for i in range(1,  (13)):
         if  ws['Q' + str(i)].value == None:
             ws['Q' + str(i)]  =  '=N' + str(i) + '-P' + str
    wb2.save(invoice) 




def null_Sales_Tax():
     for i in range(1,  (13)):
         if  ws['S' + str(i)].value == None:
             ws['S' + str(i)]  =  '=Q' + str(i) + '*R' + str(i)
     wb2.save(invoice) 




def null_InvoiceAmount():
     for i in range(1,  (13)):
         if  ws['U' + str(i)].value == None:
             ws['U' + str(i)]  = '=Q'  + str(i)+'+S' + str(i)+  '+T' + str(i)
     wb2.save(invoice) 





def null_Comments():
    for i in range(1,  (13)):
         if  ws['AB' + str(i)].value == None:
             ws['AB' + str(i)]  = Comments()
    wb2.save(invoice) 
    


def PORequestDate():

    day = random.randint(10,31)
    month = random.randint(1,9)
    year = random.randint(2020,2020)
    b  = str(day)+"/"+"0"+str(month)+"/"+str(year)
    return  b
        
        
     
 
    
def Item():
    a =("IT-279836","IT-432432","IT-63597")
    var =  random.choice(a)
    return var




def  SupplierNum():
    ca = "SU-354621","SU-324324","SU-396370","SU-354820"
    n = random.choice(ca)
    return n 



def UnitPrice():
    nu = random.uniform(10,100)
    nor =str(round(nu / 100,1)) 
    c = nor
    a = c.replace(".",",")
    return a




def DiscountRate():
    nu = random.uniform(200,1000)
    nor =str(round(nu / 100,1)) +str( "%")
    c = nor
    a = c.replace(".",",")
    return a



def SalesTax():
    nu = random.uniform(700,900)
    nor =str(round(nu / 100,1)) +str( "%")
    c = nor
    a = c.replace(".",",")
    return a




def  Suppplier():
    ca = "Company B","Company C","Company A","Company D"
    n = random.choice(ca)
    return n
   




def  ItemDescription():
    ca = "6-32 Screw","1/2-13 x 6 Hex Bolt, 18-8 Stainless Steel","1/2-13 Hex Nut"
    n = random.choice(ca)
    return n 

 

def generate_badDate():
       for i in range(4,  (13)):
           if  ws['B' + str(i)].number_format == "dd/mm/yyyy":
               ws['B' + str(i)].number_format  = "yyy/mm/dd" 
       wb2.save(invoice) 
   
    
   
    
def generate_goodDate():
       for i in range(1,  (13)):
           if  ws['B' + str(i)].number_format == "yyy/mm/dd" :
               ws['B' + str(i)].number_format  = "dd-mm-yyyy" 
       wb2.save(invoice) 
       
       
       
       
def generate_badDaePurchase():
    for i in range(4,  (12)):
           if  wk['A' + str(i)].number_format == "dd/mm/yyyy":
               wk['A' + str(i)].number_format  = "yyy/mm/dd" 
    wb3.save(purchase) 
  
    
  
def generate_goodDaePurchase():
    for i in range(4,  (12)):
           if  wk['A' + str(i)].number_format == "yyy/mm/dd":
               wk['A' + str(i)].number_format  = "dd/mm/yyyy"
    wb3.save(purchase) 
  
    
  
def pass_format():
    for i in range(1,  (12)):
        wk['A' + str(i)].number_format  = "dd/mm/yyyy"
    wb3.save(purchase) 


    
def pass_formatBB():
    for i in range(1,  (13)):
        ws['B' + str(i)].number_format  = "dd/mm/yyyy"
    wb2.save(invoice) 
 
    
 
def pass_formatC():
    for i in range(1,  (13)):
        ws['C' + str(i)].number_format  = "dd/mm/yyyy"
    wb2.save(invoice) 


    
def pass_formatD():
    for i in range(1,  (13)):
        ws['D' + str(i)].number_format  = "dd/mm/yyyy"
    wb2.save(invoice) 
    
    
    
def pass_formatV():
    for i in range(1,  (13)):
        ws['V' + str(i)].number_format  = "dd/mm/yyyy"
    wb2.save(invoice) 



def format_purchaseE():
      for i in range(1,  (12)):
          wk['E' + str(i)].number_format  = "dd/mm/yyyy" 
      wb3.save(purchase) 
    
    
    
def generateNull():
    for i in range(5,13):
        if  ws['B' + str(i)].value is not None  :
                   ws['B' + str(i)].value  = "" 
    wb2.save(invoice) 
    
    
    
def correctNull():
    for i in range(1,13):
        if  ws['B' + str(i)].value == None :
                   ws['B' + str(i)].value  = InvoiceDate()
    wb2.save(invoice) 
    
    
    
def generateNullPurchase():
    for i in range(5,12):
        if  wk['B' + str(i)].value is not None  :
                   wk['B' + str(i)].value  = "" 
    wb3.save(purchase) 
    
    
    
def correctNullPurchase():
    for i in range(1,12):
        if  wk['B' + str(i)].value == None   :
                   wk['B' + str(i)].value  = "PR-" + str(random.randint(400000,600000))
    wb3.save(purchase) 
    
    

def duplicate():
    for i in range(5,13):
      ws['A' + str(i)] = "IN-42136"
    wb2.save(invoice) 
    
    
    
def duplicatepurchase():
    for i in range(5,12):
      wk['B' + str(i)] = "PR-43530"
    wb3.save(purchase)
    

def removeDuplicate():
    for i in range(1,13):
        if  ws['A' + str(i)].value == "IN-42136":
             ws['A' + str(i)] = "IN-" + str(random.randint(10000,50000))
    wb2.save(invoice)
            
 
        

def removeDuplicatePurchase():
    for i in range(1,12):
        if  wk['A' + str(i)].value == "PR-43530":
             wk['A' + str(i)] = "PR-" + str(random.randint(400000,600000))
    wb3.save(purchase)    



    
    
def generate_date():
    generate = int(input("how much data do you want to generat: "))
    for i in range((13),  generate + 12):
         ws['A' + str(i)]  ="IN-" + str(random.randint(10000,50000))
         ws['B' + str(i)]  = InvoiceDate()
         ws['C' + str(i)]  = InvoiceDate()
         ws['D' + str(i)]  = InvoiceDate()
         ws['E' + str(i)]  = POBasedInvoices()
         ws['G' + str(i)]  ="PO-" + str(random.randint(40000,60000))
         ws['H' + str(i)]  = Item()
         ws['I' + str(i)]  = ItemDescription()
         ws['J' + str(i)]  =  "USD"
         ws['K' + str(i)]  = random.uniform(0.5, 1.7)
         a=str(ws['K' + str(i)])
         a.replace(".",",")
         ws['L' + str(i)]  = "Pieces"
         ws['M' + str(i)]  = random.uniform(4.000,20.000)
         b=str(ws['M' + str(i)])
         b.replace(".",",")
         ws['N' + str(i)]  =  '=M' + str(i) +'*K' + str(i)
         ws['O' + str(i)]  = DiscountRate()
         d=str(ws['O' + str(i)])
         d.replace(".",",")
         ws['P' + str(i)]  = '=N' + str(i) +'*O' + str(i)
         ws['Q' + str(i)]  = '=N' + str(i) + '-P' + str(i)
         ws['R' + str(i)]  = SalesTaxRate()
         v=str(ws['R' + str(i)])
         v.replace(".",",")
         ws['S' + str(i)]  = random.uniform(74.00,1.70000)
         ws['T' + str(i)]  = ShippingFee()
         ws['S' + str(i)]  = '=Q' + str(i) + '*R' + str(i) 
         ws['U' + str(i)]  = '=Q' + str(i) + '+S' + str(i) + '+T'+ str(i)
         ws['V' + str(i)]  = InvoiceDate()
         ws['W' + str(i)]  = InvoiceStatus()
         ws['X' + str(i)]  = InvoiceType()
         ws['Y' + str(i)]  = SourceSystem()
         ws['Z' + str(i)]  = SourceType()
         ws['AA' + str(i)] =  Catalog()
         ws['AB' + str(i)] =  Comments()
    wb2.save(invoice) 
 
   

         

def generate():
    generate = int(input("how much data do you want to generat: "))
    for i in range((12),  generate + 12):
         wk['A' + str(i)]  = PORequestDate()
         wk['B' + str(i)]  = "PR-" + str(random.randint(400000,600000))
         wk['C' + str(i)]  =  "PO"
         wk['D' + str(i)]  = "PO-" + str(random.randint(400000,600000))
         wk['E' + str(i)]  = PORequestDate()
         wk['F' + str(i)]  ="CO-" + str(random.randint(40000000,50000000))
         wk['G' + str(i)]  = "Production"
         wk['H' + str(i)]  =  SupplierNum()
         wk['I' + str(i)]  =  Suppplier()
         wk['J' + str(i)]  =   Item()
         wk['K' + str(i)]  =   ItemDescription()
         wk['L' + str(i)]  = "USD"
         wk['M' + str(i)]  = UnitPrice()
         wk['N' + str(i)]  =  "Pieces"
         wk['O' + str(i)]  =  random.uniform(5.000,29.0000)
         r = str(wk['O' + str(i)]) 
         r.replace(".",",")
         wk['P' + str(i)]  = '=O' + str(i) +'*M' + str(i)
         wk['Q' + str(i)]  =  DiscountRate()
         wk['R' + str(i)]  = '=P' + str(i) + '*Q' + str(i) 
         wk['S' + str(i)]  = SalesTax()
         wk['S' + str(i)]  = SalesTax()
         wk['T' + str(i)]  = '=(P' + str(i) +'-R'+str(i) +')*(1+S' + str(i) +')'
         wk['U' + str(i)]  = "Quantity based volume discount"
        
    wb3.save(purchase) 



value = True
while value == True:
    os.system("cls")

    print("\n+--------------------------------------------+")
    print("|            SELECT AT FILE                    |")
    print("|           ----------------------             |")
    print("|       1.Invoice.xlsx                         |")
    print("|       2.Purchase Order.xlsx                  |")
    print("+--------------------------------------------+")

    option=int(input("choose an option: "))
    if option ==1:
        invoice = os.getcwd()+"\\Invoice.xlsx"
        date = pd.read_excel(invoice, sheet_name='Invoice', engine='openpyxl')
        wb2 = Workbook()
        wb2 = load_workbook(invoice)
        ws = wb2.active
        col=[]
        col = date.columns.values
        desired_columns=[]
        desired_columns = date.columns.values
        
        value = False

        
    elif option ==2:
        purchase = os.getcwd()+"\\Purchase Order.xlsx"
        df = pd.read_excel(purchase, sheet_name='PurchaseOrder', engine='openpyxl')
        wb3 = Workbook()
        wb3 = load_workbook(purchase)
        wk = wb3.active
        col=[]
        col = df.columns.values
        desired_columns=[]
        desired_columns = df.columns.values
        value = False


fact = True
while (fact == True):
            os.system("pause")
            os.system("cls")
            print ("\n+----------------------------------------------------+")
            print("             ADMIN OF DATE FILE:                        " )
            print("|           ----------------------                      |")
            print("|         1.show the columns you want to see            |")
            print("|         2.show single column                          |")
            print("|         3.show number of rows                         |")
            print("|         4.show single of rows                         |")
            print("|         5.show how many empty elements in each column |")
            print("|         6.show column attributes                      |")
            print("|         7.show all coolumn atributes                  |")
            print("|         8.you show number of rows                     |")
            print("|         9.you show number of column                   |")
            print("|         10.you show number of non-null elements       |")
            print("|         11.dataframe size                             |")
            print("|         12.replace empty values                       |")
            print("|         13.generate data in file                      |")
            print("|         14.pass dates to format dd-mm-yyyy            |")
            print("|         15.generate bad date yyy-mm-dd                |")
            print("|         16.date yyy-mm-dd at dd-mm-yyyy               |")
            print("|         17.add empty date                             |")
            print("|         18.correct empty data                         |")
            print("|         19.add duplicate data                         |")
            print("|         20.remove duplicate data                      |")
            print("|         0.exit                                        |")
            print("+-------------------------------------------------------+")
            

            val= input("welcome dear teache, please choose an option: ")
            
            if val == "1":
                if option ==1:
                    viewColumns()
                elif option ==2:
                    viewColumns_purchase()
            elif val == "2":
                 if option ==1:
                     viewColumns_voice()
                 elif option ==2:
                     columns_purchase()
            elif val == "3":
                if option ==1:
                     row_invoice()
                elif option ==2:
                    row_purchase()
            elif val == "4":
                if option ==1:
                    row_single()
                elif option ==2:
                    row_single_purchase()
            elif val == "5":
                 if option ==1:
                     empty()
                 elif option ==2:
                     empty_purchase()
            elif val == "6":
                 if option ==1:
                     attributes()
                 elif option ==2:
                     attributes_purchase()
            elif val == "7":
                 if option ==1:
                     show_all()
                 elif option ==2:
                     show_all_purchase()
            elif val == "8":
                 if option ==1:
                     show_rows()
                 elif option ==2:
                     show_rows_purchase()
            elif val == "9":
                 if option ==1:
                      show_column()
                 elif option ==2:
                     show_column_purchase()
            elif val == "10":
                 if option ==1:
                      not_null()
                 elif option ==2:
                     not_null_purchase()
            elif val == "11":
                 if option ==1:
                      zise()
                 elif option ==2:
                      size_purchase()
            elif val == "12":
                 if option ==1:
                      null()
                      null_Comments()
                 elif option == 2:
                    print("there are no empty elements")
            elif val == "13":
                 if option ==1:
                      generate_date()
                 elif option ==2:
                     generate()
            elif val == "14":
                 if option ==1:
                      pass_formatC()
                      pass_formatBB()
                      pass_formatD()
                      pass_formatV()
                 elif option ==2:
                      pass_format()
                      format_purchaseE()
            elif val == "15":
                 if option ==1:
                     generate_badDate()
                 elif option ==2:
                     generate_badDaePurchase()
            elif val == "16":
                 if option ==1:
                     generate_goodDate()
                 elif option ==2:
                     generate_goodDaePurchase()
            elif val == "17":
                 if option ==1:
                   generateNull()
                 elif option ==2:
                  generateNullPurchase()
            elif val == "18":
                 if option ==1:
                   correctNull()
                 elif option ==2:
                   correctNullPurchase()
            elif val == "19":
                 if option ==1:
                   duplicate()
                 elif option ==2:
                   duplicatepurchase()
            elif val == "20":
                 if option ==1:
                   removeDuplicate()
                 elif option ==2:
                   duplicatepurchase()
            elif val == "0":
                fact = False
